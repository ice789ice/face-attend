import cv2
import os
import numpy as np
import pandas as pd
from concurrent.futures import ProcessPoolExecutor, as_completed
from tqdm import tqdm
import time
import logging
import logging.handlers
import multiprocessing
from functools import partial
import sys
from openpyxl import load_workbook

# ================= 配置类 =================
class Config:
    # 基础路径配置
    BASE_DIR = r"D:\face-attend\WorkAttendanceSystem-master\V2.0"
    TEST_IMAGES_DIR = os.path.join(BASE_DIR, "test_images")
    RESULTS_DIR = os.path.join(BASE_DIR, "results")
    EXCEL_REPORT = os.path.join(BASE_DIR, "detection_report.xlsx")
    LOG_FILE = os.path.join(BASE_DIR, "face_detection.log")
    
    # 模型选择 (dnn/haar)
    MODEL_TYPE = 'dnn'
    
    # 性能参数
    MAX_IMAGES = 20000  # 只处理前2万张
    BATCH_SIZE = 500    # 减小批次大小
    PROCESS_WORKERS = min(8, max(1, multiprocessing.cpu_count() - 2))  # 限制最大进程数
    
    # DNN参数
    DNN_CONFIDENCE = 0.7  # 提高置信度阈值
    DNN_INPUT_SIZE = (300, 300)
    DNN_SCALE_FACTOR = 1.0
    DNN_MEAN_VALUES = (104.0, 177.0, 123.0)
    
    # Haar参数
    HAAR_SCALE_FACTOR = 1.05
    HAAR_MIN_NEIGHBORS = 6
    HAAR_MIN_SIZE = (30, 30)

# ================= 初始化日志系统 =================
def setup_logging():
    """配置多级别日志输出"""
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)
    
    # 清除现有handler
    for handler in logger.handlers[:]:
        logger.removeHandler(handler)
    
    # 文件日志（带滚动）
    file_handler = logging.handlers.RotatingFileHandler(
        Config.LOG_FILE, 
        maxBytes=10*1024*1024, 
        backupCount=5,
        encoding='utf-8'
    )
    file_handler.setFormatter(logging.Formatter('%(asctime)s [%(levelname)s] %(message)s'))
    
    # 控制台日志
    console_handler = logging.StreamHandler()
    console_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
    
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)
    logging.info("日志系统初始化完成")

# ================= 模型加载器 =================
class ModelLoader:
    @staticmethod
    def load_dnn_model():
        """加载DNN模型"""
        proto_path = os.path.join(Config.BASE_DIR, "deploy.prototxt")
        model_path = os.path.join(Config.BASE_DIR, "res10_300x300_ssd_iter_140000.caffemodel")
        
        if not all(os.path.exists(p) for p in [proto_path, model_path]):
            raise FileNotFoundError("DNN模型文件缺失")
            
        net = cv2.dnn.readNetFromCaffe(proto_path, model_path)
        
        if cv2.cuda.getCudaEnabledDeviceCount() > 0:
            net.setPreferableBackend(cv2.dnn.DNN_BACKEND_CUDA)
            net.setPreferableTarget(cv2.dnn.DNN_TARGET_CUDA)
            logging.info("使用CUDA加速")
        else:
            net.setPreferableBackend(cv2.dnn.DNN_BACKEND_OPENCV)
            net.setPreferableTarget(cv2.dnn.DNN_TARGET_CPU)
            logging.info("使用CPU模式")
            
        return net

    @staticmethod
    def load_haar_model():
        """加载Haar级联分类器"""
        model_path = cv2.data.haarcascades + 'haarcascade_frontalface_default.xml'
        if not os.path.exists(model_path):
            raise FileNotFoundError(f"Haar模型文件不存在: {model_path}")
        return cv2.CascadeClassifier(model_path)

# ================= 人脸检测器 =================
class FaceDetector:
    @staticmethod
    def dnn_detect(img, net):
        """DNN人脸检测"""
        (h, w) = img.shape[:2]
        blob = cv2.dnn.blobFromImage(
            cv2.resize(img, Config.DNN_INPUT_SIZE), 
            Config.DNN_SCALE_FACTOR, 
            Config.DNN_INPUT_SIZE, 
            Config.DNN_MEAN_VALUES
        )
        net.setInput(blob)
        detections = net.forward()
        
        faces = []
        for i in range(detections.shape[2]):
            confidence = detections[0, 0, i, 2]
            if confidence > Config.DNN_CONFIDENCE:
                box = detections[0, 0, i, 3:7] * np.array([w, h, w, h])
                faces.append(box.astype("int"))
        return faces

    @staticmethod
    def haar_detect(img, classifier):
        """Haar人脸检测"""
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        return classifier.detectMultiScale(
            gray,
            scaleFactor=Config.HAAR_SCALE_FACTOR,
            minNeighbors=Config.HAAR_MIN_NEIGHBORS,
            minSize=Config.HAAR_MIN_SIZE
        )

# ================= 单图片处理器 =================
def process_single_image(img_path, model_type):
    """处理单张图片"""
    try:
        # 每个进程独立加载模型
        if model_type == 'dnn':
            model = ModelLoader.load_dnn_model()
            detector = FaceDetector.dnn_detect
        else:
            model = ModelLoader.load_haar_model()
            detector = FaceDetector.haar_detect
        
        # 读取图片
        img = cv2.imread(img_path)
        if img is None:
            return os.path.basename(img_path), 0, "读取失败"
        
        # 检测人脸
        faces = detector(img, model)
        
        # 保存结果图
        output_path = os.path.join(Config.RESULTS_DIR, f"detected_{os.path.basename(img_path)}")
        for (x, y, w, h) in faces:
            cv2.rectangle(img, (x, y), (x + w, y + h), (0, 255, 0), 2)
        cv2.imwrite(output_path, img)
        
        return os.path.basename(img_path), len(faces), "成功"
        
    except Exception as e:
        return os.path.basename(img_path), 0, f"错误: {str(e)}"

# ================= 批处理器 =================
class BatchProcessor:
    @staticmethod
    def run_batch(batch_files, batch_num):
        """处理单个批次"""
        logging.info(f"开始处理批次 {batch_num} (共 {len(batch_files)} 张图片)")
        
        processor = partial(process_single_image, model_type=Config.MODEL_TYPE)
        results = []
        
        with ProcessPoolExecutor(max_workers=Config.PROCESS_WORKERS) as executor:
            # 创建任务映射
            future_to_file = {
                executor.submit(processor, img): os.path.basename(img)
                for img in batch_files
            }
            
            # 处理完成的任务
            for future in tqdm(
                as_completed(future_to_file),
                total=len(batch_files),
                desc=f"批次 {batch_num}",
                mininterval=5  # 减少进度条更新频率
            ):
                try:
                    results.append(future.result())
                except Exception as e:
                    file_name = future_to_file[future]
                    logging.error(f"处理 {file_name} 失败: {str(e)}")
                    results.append((file_name, 0, f"错误: {str(e)}"))
        
        return results

    @staticmethod
    def save_results(results, batch_num):
        """安全保存批次结果"""
        df = pd.DataFrame(results, columns=["图片名", "人脸数", "状态"])
        csv_path = Config.EXCEL_REPORT.replace('.xlsx', '.csv')
        
        try:
            # 首次写入创建新文件
            if batch_num == 0:
                df.to_excel(Config.EXCEL_REPORT, index=False, engine='openpyxl')
                df.to_csv(csv_path, index=False)
                logging.info(f"创建新的结果文件: {Config.EXCEL_REPORT}")
            else:
                # 追加到Excel
                book = load_workbook(Config.EXCEL_REPORT)
                writer = pd.ExcelWriter(Config.EXCEL_REPORT, engine='openpyxl')
                writer.book = book
                writer.sheets = {ws.title: ws for ws in book.worksheets}
                
                startrow = writer.sheets['Sheet1'].max_row
                df.to_excel(
                    writer,
                    startrow=startrow,
                    header=False,
                    index=False
                )
                writer.close()
                
                # 追加到CSV
                df.to_csv(csv_path, mode='a', header=False, index=False)
                
            logging.info(f"批次 {batch_num} 完成，成功 {len(df[df['状态']=='成功'])} 张")
            
        except Exception as e:
            logging.error(f"保存结果失败: {str(e)}，仅保存到CSV")
            df.to_csv(csv_path, mode='a', header=not os.path.exists(csv_path), index=False)

# ================= 主控制器 =================
def main_controller():
    """主控制流程"""
    setup_logging()
    os.makedirs(Config.RESULTS_DIR, exist_ok=True)
    
    try:
        # 获取图片列表（只处理前2万张）
        all_files = [
            os.path.join(Config.TEST_IMAGES_DIR, f) 
            for f in sorted(os.listdir(Config.TEST_IMAGES_DIR))
            if f.lower().endswith(('.jpg', '.jpeg', '.png'))
        ]
        image_files = all_files[:Config.MAX_IMAGES]
        
        if not image_files:
            logging.error("未找到有效图片文件！")
            return
        
        logging.info(f"\n{'='*50}\n启动人脸检测系统\n"
                    f"模型类型: {Config.MODEL_TYPE.upper()}\n"
                    f"处理图片数: {len(image_files)}/{len(all_files)}\n"
                    f"工作进程: {Config.PROCESS_WORKERS}\n"
                    f"批处理量: {Config.BATCH_SIZE}\n"
                    f"{'='*50}")
        
        # 分批处理
        start_time = time.time()
        total_batches = (len(image_files) + Config.BATCH_SIZE - 1) // Config.BATCH_SIZE
        
        for batch_num, i in enumerate(range(0, len(image_files), Config.BATCH_SIZE)):
            batch = image_files[i:i + Config.BATCH_SIZE]
            logging.info(f"处理批次 {batch_num+1}/{total_batches} (图片 {i+1}-{min(i+Config.BATCH_SIZE, len(image_files))})")
            
            results = BatchProcessor.run_batch(batch, batch_num)
            BatchProcessor.save_results(results, batch_num)
            
            # 每5批释放一次内存
            if batch_num % 5 == 0:
                import gc
                gc.collect()
        
        # 生成最终报告
        generate_final_report(start_time, len(image_files))
        
    except Exception as e:
        logging.exception("主流程发生错误")
    finally:
        logging.info("处理结束")

def generate_final_report(start_time, total_images):
    """生成汇总报告"""
    try:
        csv_path = Config.EXCEL_REPORT.replace('.xlsx', '.csv')
        if os.path.exists(csv_path):
            df = pd.read_csv(csv_path)
        else:
            df = pd.read_excel(Config.EXCEL_REPORT)
        
        elapsed = time.time() - start_time
        
        stats = {
            "处理日期": time.strftime("%Y-%m-%d %H:%M:%S"),
            "总图片数": total_images,
            "成功数": len(df[df["状态"] == "成功"]),
            "失败数": len(df[df["状态"] != "成功"]),
            "总人脸数": int(df["人脸数"].sum()),
            "平均人脸数": round(df["人脸数"].mean(), 2),
            "总耗时": f"{elapsed//3600:.0f}小时 {elapsed%3600//60:.0f}分钟",
            "处理速度(图片/秒)": round(total_images/elapsed, 2),
            "失败原因统计": str(df[df["状态"] != "成功"]["状态"].value_counts().to_dict())
        }
        
        # 保存统计结果
        stats_df = pd.DataFrame.from_dict(stats, orient='index', columns=['数值'])
        with pd.ExcelWriter(Config.EXCEL_REPORT, engine='openpyxl', mode='a') as writer:
            stats_df.to_excel(writer, sheet_name='统计摘要')
        
        # 打印摘要
        logging.info(f"\n{'='*50}\n处理完成！\n"
                    f"总图片数: {stats['总图片数']}\n"
                    f"成功数: {stats['成功数']} ({stats['成功数']/stats['总图片数']*100:.2f}%)\n"
                    f"总人脸数: {stats['总人脸数']}\n"
                    f"总耗时: {stats['总耗时']}\n"
                    f"处理速度: {stats['处理速度(图片/秒)']} 图片/秒\n"
                    f"{'='*50}")
        
        # 保存统计到CSV
        stats_df.to_csv(csv_path.replace('.csv', '_stats.csv'))
        
    except Exception as e:
        logging.error(f"生成报告失败: {str(e)}")

if __name__ == "__main__":
    # Windows多进程保护
    multiprocessing.freeze_support()
    
    # 版本检查
    if sys.version_info < (3, 7):
        print("需要Python 3.7或更高版本")
        sys.exit(1)
    
    # 依赖检查
    try:
        import openpyxl
    except ImportError:
        logging.warning("未找到openpyxl，Excel报告功能将受限")
    
    main_controller()